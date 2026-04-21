# -*- coding: utf-8 -*-
"""
Data Mining Project 2
Authors : Miles Glover, Madison Nicholson, Victory Orobosa

Requirements:
    pip install numpy matplotlib scikit-learn scikit-fuzzy openpyxl
"""

import os
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import random
from openpyxl import load_workbook
from sklearn.preprocessing import MinMaxScaler
import skfuzzy as fuzz

# ============================================================
# 1. Load Dataset
# ============================================================

def load_XLSX(file_name):
    """Load the .xlsx gene expression dataset."""
    print("Loading file...")

    dataset = []

    wb = load_workbook(file_name)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])

    for row in rows[1:]:
        dataset.append(list(row))

    print(f"File loaded: {len(dataset)} rows, {len(header)} columns\n")
    return dataset, header


# ============================================================
# 2. Data Understanding (Victory)
# ============================================================

def data_understanding(data, header):
    """
    Print a summary of the dataset: column names, observation count,
    and basic statistics for the three numeric features.
    """
    print("=" * 60)
    print("DATASET UNDERSTANDING")
    print("=" * 60)
    print(f"\nDataset: Longotor1delta (Yeast Gene Expression)")
    print(f"Number of observations: {len(data)}")
    print(f"Number of columns: {len(header)}")
    print(f"\nColumn names: {header}")

    print("\nColumn Descriptions:")
    print("  - Public ID       : Systematic yeast ORF identifier (e.g. YAL001C)")
    print("  - Gene            : Standard gene name")
    print("  - Gene description: Functional annotation of the gene")
    print("  - sch9/wt         : Log-ratio expression (sch9 knockout vs wild type)")
    print("  - ras2/wt         : Log-ratio expression (ras2 knockout vs wild type)")
    print("  - tor1/wt         : Log-ratio expression (tor1 knockout vs wild type)")
    print("  - Columns 7-9     : Empty columns (all NaN) -- will be dropped\n")

    # Quick numeric stats on the 3 feature columns (indices 3,4,5)
    features = np.array([[row[3], row[4], row[5]] for row in data], dtype=float)
    feature_names = [header[3], header[4], header[5]]
    for i, name in enumerate(feature_names):
        col = features[:, i]
        print(f"  {name:>10s}  |  min={col.min():.4f}  max={col.max():.4f}  "
              f"mean={col.mean():.4f}  std={col.std():.4f}")

    print()
    return features, feature_names


# ============================================================
# 3. Preprocessing and Normalization (Victory)
# ============================================================

def preprocess(data, header):
    """
    Select the three numeric feature columns, drop unused columns,
    and return a clean numpy array along with gene labels.
    """
    print("Preprocessing: selecting numeric columns (sch9/wt, ras2/wt, tor1/wt)...")

    features = np.array([[row[3], row[4], row[5]] for row in data], dtype=float)
    gene_ids = [row[0] for row in data]       # Public ID for cluster membership
    gene_names = [row[1] for row in data]     # Gene name

    # Check for any NaN in the 3 feature columns
    nan_count = np.isnan(features).sum()
    if nan_count > 0:
        print(f"  Warning: {nan_count} NaN values found in feature columns. Rows with NaN will be removed.")
        mask = ~np.isnan(features).any(axis=1)
        features = features[mask]
        gene_ids = [g for g, m in zip(gene_ids, mask) if m]
        gene_names = [g for g, m in zip(gene_names, mask) if m]
    else:
        print("  No missing values in the 3 feature columns.")

    print(f"  Final dataset shape: {features.shape}\n")
    return features, gene_ids, gene_names


def normalize(features, feature_names):
    """
    Apply Min-Max normalization to scale all feature values to [0, 1].
    Returns the normalized features and the fitted scaler.
    """
    print("Normalizing features using Min-Max scaling (0 to 1)...")

    scaler = MinMaxScaler()
    features_norm = scaler.fit_transform(features)

    print("  Before normalization:")
    for i, name in enumerate(feature_names):
        col = features[:, i]
        print(f"    {name:>10s}  min={col.min():.4f}  max={col.max():.4f}")

    print("  After normalization:")
    for i, name in enumerate(feature_names):
        col = features_norm[:, i]
        print(f"    {name:>10s}  min={col.min():.4f}  max={col.max():.4f}")

    print()
    return features_norm, scaler


# ============================================================
# 4. Fuzzy C-Means Clustering (Victory)
# ============================================================

def fuzzy_cmeans_clustering(features_norm, n_clusters=3, m=2.0, error=0.005, maxiter=1000):
    """
    Perform Fuzzy C-Means clustering.

    Parameters
    ----------
    features_norm : np.ndarray, shape (n_samples, n_features)
        Normalized feature matrix.
    n_clusters : int
        Number of clusters (c).
    m : float
        Fuzziness parameter (m > 1). Higher m = softer/fuzzier clusters.
    error : float
        Stopping criterion (minimum improvement in objective function).
    maxiter : int
        Maximum number of iterations.

    Returns
    -------
    cntr : np.ndarray
        Cluster centers.
    u : np.ndarray
        Final membership matrix (n_clusters x n_samples).
    labels : np.ndarray
        Hard cluster labels based on max membership.
    fpc : float
        Fuzzy Partition Coefficient (1 = crisp, 1/c = completely fuzzy).
    """
    print(f"Running Fuzzy C-Means: c={n_clusters}, m={m}, error={error}, maxiter={maxiter}")

    # skfuzzy expects data as (n_features, n_samples), so transpose
    data_T = features_norm.T

    cntr, u, u0, d, jm, p, fpc = fuzz.cluster.cmeans(
        data_T,
        c=n_clusters,
        m=m,
        error=error,
        maxiter=maxiter,
        seed=int(random.random()*10000)
    )

    # Hard labels: assign each point to the cluster with highest membership
    labels = np.argmax(u, axis=0)

    print(f"  Converged in {p} iterations")
    print(f"  Fuzzy Partition Coefficient (FPC): {fpc:.4f}")
    for i in range(n_clusters):
        count = int(np.sum(labels == i))
        print(f"  Cluster {i}: {count} members")
    print()

    return cntr, u, labels, fpc


# ============================================================
# 5. FCM Evaluation & Visualization (Victory)
# ============================================================

def plot_fcm_clusters(features_norm, labels, cntr, feature_names, n_clusters, m, save_dir="plots", save_prefix="fcm"):
    """Create 2D scatter plots of FCM cluster assignments for each pair of features."""
    pairs = [(0, 1), (0, 2), (1, 2)]
    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    fig.suptitle(f"Fuzzy C-Means Clustering (c={n_clusters}, m={m})", fontsize=14)

    colors = plt.cm.tab10(np.linspace(0, 1, n_clusters))

    for ax, (xi, yi) in zip(axes, pairs):
        for k in range(n_clusters):
            mask = labels == k
            ax.scatter(features_norm[mask, xi], features_norm[mask, yi],
                       c=[colors[k]], alpha=0.4, s=10, label=f"Cluster {k}")
        # Plot cluster centers
        ax.scatter(cntr[:, xi], cntr[:, yi], c='black', marker='X', s=150,
                   edgecolors='white', linewidths=1.5, label='Centers')
        ax.set_xlabel(feature_names[xi])
        ax.set_ylabel(feature_names[yi])
        ax.legend(fontsize=8)

    plt.tight_layout()
    fname = os.path.join(save_dir, f"{save_prefix}_c{n_clusters}_m{m}.png")
    plt.savefig(fname, dpi=150)
    plt.close()
    print(f"  Saved plot: {fname}")


def plot_membership_heatmap(u, n_clusters, m, save_dir="plots", save_prefix="fcm"):
    """Show a heatmap of membership values for a sample of points."""
    n_show = min(50, u.shape[1])
    fig, ax = plt.subplots(figsize=(14, 4))
    im = ax.imshow(u[:, :n_show], aspect='auto', cmap='YlOrRd')
    ax.set_xlabel("Sample Index")
    ax.set_ylabel("Cluster")
    ax.set_yticks(range(n_clusters))
    ax.set_title(f"FCM Membership Matrix (first {n_show} samples, c={n_clusters}, m={m})")
    plt.colorbar(im, ax=ax, label="Membership Degree")
    plt.tight_layout()
    fname = os.path.join(save_dir, f"{save_prefix}_membership_c{n_clusters}_m{m}.png")
    plt.savefig(fname, dpi=150)
    plt.close()
    print(f"  Saved plot: {fname}")


def print_cluster_members(labels, gene_ids, gene_names, n_clusters, max_per_cluster=10):
    """Print a sample of gene members from each cluster."""
    print(f"\nCluster Membership Sample (up to {max_per_cluster} per cluster):")
    print("-" * 60)
    for k in range(n_clusters):
        indices = np.where(labels == k)[0]
        print(f"\nCluster {k} ({len(indices)} total members):")
        for idx in indices[:max_per_cluster]:
            print(f"  {gene_ids[idx]:>12s}  {gene_names[idx]}")
        if len(indices) > max_per_cluster:
            print(f"  ... and {len(indices) - max_per_cluster} more")


def evaluate_fpc_vs_clusters(features_norm, m=2.0, k_range=range(2, 11), save_dir="plots"):
    """
    Run FCM for different numbers of clusters and plot the
    Fuzzy Partition Coefficient to help choose the best k.
    """
    print("Evaluating FPC across different cluster counts...")
    fpcs = []
    for c in k_range:
        _, u, _, _, _, _, fpc = fuzz.cluster.cmeans(
            features_norm.T, c=c, m=m, error=0.005, maxiter=1000, seed=int(random.random()*10000)
        )
        fpcs.append(fpc)
        print(f"  c={c}: FPC={fpc:.4f}")

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(list(k_range), fpcs, 'bo-', linewidth=2)
    ax.set_xlabel("Number of Clusters (c)")
    ax.set_ylabel("Fuzzy Partition Coefficient (FPC)")
    ax.set_title(f"FPC vs Number of Clusters (m={m})")
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    fname = os.path.join(save_dir, "fcm_fpc_vs_clusters.png")
    plt.savefig(fname, dpi=150)
    plt.close()
    print(f"  Saved plot: {fname}\n")

    return fpcs


def evaluate_fuzziness_effect(features_norm, n_clusters=3, m_values=[1.5, 2.0, 2.5, 3.0]):
    """
    Run FCM with different fuzziness parameters to see
    how m affects cluster crispness.
    """
    print("Evaluating the effect of fuzziness parameter m...")
    results = []
    for m in m_values:
        cntr, u, labels, fpc = fuzzy_cmeans_clustering(features_norm, n_clusters=n_clusters, m=m)
        results.append((m, cntr, u, labels, fpc))
    return results


# ============================================================
# K-Means (Miles Glover)
# ============================================================

# helper distance function to calculate either the l1 or l2
# distance between two data points (data rows)
def distance(data_row_1, data_row_2, distance_method):

    total_dist = 0
    difference = 0
    exp = 1
    
    if distance_method == "L1":
        exp = 1
    if distance_method == "L2":
        exp = 2

    for i in range(len(data_row_1)):

        difference = data_row_1[i] - data_row_2[i]
        difference = pow(abs(difference), exp)
        total_dist += difference

    return pow(total_dist, 1/exp)

# Function that performs k means 
# clustering on a datased with a specified distance function,
# k value, iteration limit, and convergence threshhold
def k_means_clustering(k, iter_lim, threshold, distance_method, data):

    #function specific variables    

    iterations = 0
    cluster_positions = []
    cluster_assignment = []
    data_dimensions = len(data[0])
    data_rows = len(data)
    pre_error = 0

    # [step 0] set initial cluster positions (random position between [-1,1] per attribute)
    
    random.seed(int(random.random()*10000))
    
    for i in range(k):

        cluster = []

        for j in range(data_dimensions):

            cluster.append((random.random()*2)-1)

        cluster_positions.append(cluster)

    # iterate
    while iterations < iter_lim:

        error = 0
        cluster_assignment = []

        # [step 1] assign each data point to its closest cluster center
        for row in range(data_rows):

            for clust in range(len(cluster_positions)):

                dist = distance(cluster_positions[clust], data[row], distance_method)

                if clust == 0:

                    least_distant = dist, clust

                if least_distant[0] > dist:

                    least_distant = dist, clust

            cluster_assignment.append(least_distant)

            if distance_method == "L2":
                
                error += pow(least_distant[0], 2)
                
            elif distance_method == "L1":
                
                error += least_distant[0]


        # [step 2] recalculate cluster centers

        assignment_count = []
        cluster_positions = []

        for i in range(k):

            cluster = []

            for j in range(data_dimensions):

                cluster.append(0)

            cluster_positions.append(cluster)
            assignment_count.append(0)

        # sum data values into each cluster
        for row in range(data_rows):

            dist, clust = cluster_assignment[row]
            assignment_count[clust] += 1

            for col in range(data_dimensions):

                cluster_positions[clust][col] += data[row][col]

        # update cluster positions
        for clust in range(k):

            if assignment_count[clust] > 0:

                for col in range(data_dimensions):

                    # average position (L2)
                    if distance_method == "L2":

                        cluster_positions[clust][col] = cluster_positions[clust][col] / assignment_count[clust]

                    # median position (L1)
                    elif distance_method == "L1":

                        cluster_values = []

                        for row in range(data_rows):

                            dist, assigned_clust = cluster_assignment[row]

                            if assigned_clust == clust:

                                cluster_values.append(data[row][col])

                        cluster_values.sort()
                        middle = int(len(cluster_values) / 2)

                        if len(cluster_values) % 2 == 0:
                            
                            cluster_positions[clust][col] = (cluster_values[middle - 1] + cluster_values[middle]) / 2
                        else:
                            
                            cluster_positions[clust][col] = cluster_values[middle]

        # exit condition (convergence check)
        if iterations != 0:
            
            if pre_error != 0 and ((pre_error - error) / pre_error) < threshold:
                
                break

        pre_error = error
        iterations += 1

    return cluster_positions, cluster_assignment, abs(error)

# ============================================================
# K-Means Evaluation & Visualization (Madison)
# ============================================================

# creates scatter plots of K-Means cluster assignments
def plot_kmeans_clusters(features_norm, cluster_assignment, cluster_positions, feature_names,
                         k, dist_method, save_dir="plots"):
    labels = np.array([ca[1] for ca in cluster_assignment])
    centers = np.array(cluster_positions)

    pairs = [(0, 1), (0, 2), (1, 2)]
    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    fig.suptitle(f"K-Means Clustering (k={k}, distance={dist_method})", fontsize=14)

    colors = plt.cm.tab10(np.linspace(0, 1, k))

    for ax, (xi, yi) in zip(axes, pairs):
        for c in range(k):
            mask = labels == c
            ax.scatter(features_norm[mask, xi], features_norm[mask, yi],
                       c=[colors[c]], alpha=0.4, s=10, label=f"Cluster {c}")
        ax.scatter(centers[:, xi], centers[:, yi], c='black', marker='X', s=150,
                   edgecolors='white', linewidths=1.5, label='Centers')
        ax.set_xlabel(feature_names[xi])
        ax.set_ylabel(feature_names[yi])
        ax.legend(fontsize=8)

    plt.tight_layout()
    fname = os.path.join(save_dir, f"kmeans_k{k}_{dist_method}.png")
    plt.savefig(fname, dpi=150)
    plt.close()
    print(f"  Saved plot: {fname}")

#runs kmeans for different k and creates elbow plot to show error curve
def evaluate_kmeans_sse(features_norm_list, k_range=range(2, 11), dist_method="L2", save_dir="plots"):

    print(f"Evaluating K-Means SSE across cluster counts (distance={dist_method})...")
    errors = []
    for k in k_range:
        _, _, err = k_means_clustering(k, 100, 0.001, dist_method, features_norm_list)
        errors.append(err)
        print(f"  k={k}: SSE={err:.4f}")

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(list(k_range), errors, 'ro-', linewidth=2)
    ax.set_xlabel("Number of Clusters (k)")
    ax.set_ylabel("Sum of Squared Errors (SSE)")
    ax.set_title(f"Elbow Method - K-Means ({dist_method})")
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    fname = os.path.join(save_dir, f"kmeans_elbow_{dist_method}.png")
    plt.savefig(fname, dpi=150)
    plt.close()
    print(f"  Saved plot: {fname}\n")

    return errors

#identifies members of each cluster & analyzes whether clusters reveal good groupings
def print_kmeans_cluster_members(cluster_assignment, gene_ids, gene_names, k, max_per_cluster=10):
    """Print a sample of gene members from each K-Means cluster."""
    labels = [ca[1] for ca in cluster_assignment]
    print(f"\nK-Means Cluster Membership Sample (up to {max_per_cluster} per cluster):")
    print("-" * 60)
    for c in range(k):
        indices = [i for i, l in enumerate(labels) if l == c]
        print(f"\nCluster {c} ({len(indices)} total members):")
        for idx in indices[:max_per_cluster]:
            print(f"  {gene_ids[idx]:>12s}  {gene_names[idx]}")
        if len(indices) > max_per_cluster:
            print(f"  ... and {len(indices) - max_per_cluster} more")

# =========================================================
# COMPARISON ANALYSIS (Madison)
# =========================================================
from sklearn.metrics import silhouette_score, davies_bouldin_score, calinski_harabasz_score

sil_score = silhouette_score
db_score = davies_bouldin_score
ch_score = calinski_harabasz_score

#kmeans - all 3 internal validation 
def internal_kmeans(features_norm, cluster_assignment, k):
    labels = np.array([ca[1] for ca in cluster_assignment])
    
    unique_labels = np.unique(labels)
    if len(unique_labels) < 2:
        return None
        
    results = {
       "calinski_harabasz": ch_score(features_norm, labels),
        "silhouette": sil_score(features_norm, labels),
        "davies_bouldin": db_score(features_norm, labels)
    }
    
    #printing results to 4 decimal places
    print(f"\nKMeans' Internal Validation Methods for k = {k}: ")
    print(f" Calinski-Harabasz Score: {results['calinski_harabasz']:.4f}")
    print(f" Davies-Bouldin Score: {results['davies_bouldin']:.4f}")
    print(f" Silhouette Score: {results['silhouette']:.4f}")
    
    return results
    
#fcm - all 3 internal validation 
def internal_fcm(features_norm, u, k):
    labels = np.argmax(u, axis=0)
    
    unique_labels = np.unique(labels)
    if len(unique_labels) < 2:
        return None
        
    results = {
       "calinski_harabasz": ch_score(features_norm, labels),
        "silhouette": sil_score(features_norm, labels),
        "davies_bouldin": db_score(features_norm, labels)
    }
    
    #printing results to 4 decimal places
    print(f"\nFCM's Internal Validation Methods for k = {k}:")
    print(f" Calinski-Harabasz Score: {results['calinski_harabasz']:.4f}")
    print(f" Davies-Bouldin Score: {results['davies_bouldin']:.4f}")
    print(f" Silhouette Score: {results['silhouette']:.4f}")
    
    return results
                
#plotting the results in a line graph -  Kmeans vs FCM Internal Metrics
    # x-axis = number of clusters
    # y-axis = internal validation metric value
def plot_internal(kmeans_results, fcm_results, save_dir="plots"):
    
    #k values per model
    kmeans_k = [r["k"] for r in kmeans_results]
    fcm_k = [r["k"] for r in fcm_results]
    
    #plotting silhouette
    plt.figure()
    plt.plot(kmeans_k, [r["silhouette"] for r in kmeans_results], marker='o', label="KMeans")
    plt.plot(fcm_k, [r["silhouette"] for r in fcm_results], marker='s', label="FCM")
    plt.xlabel("Cluster Amount (k)")
    plt.ylabel("Silhouette Score")
    plt.title("Silhouette Score Comparison")
    plt.legend()
    plt.grid(True, alpha=0.3) #formatting
    plt.savefig(os.path.join(save_dir, "silhouette_comparison.png"))
    plt.close()
    
    #plotting DB
    plt.figure()
    plt.plot(kmeans_k, [r["davies_bouldin"] for r in kmeans_results], marker='o', label="KMeans")
    plt.plot(fcm_k, [r["davies_bouldin"] for r in fcm_results], marker='s', label="FCM")
    plt.xlabel("Cluster Amount (k)")
    plt.ylabel("Davies-Bouldin Score")
    plt.title("Davies-Bouldin Score Comparison")
    plt.legend()
    plt.grid(True, alpha=0.3) #formatting
    plt.savefig(os.path.join(save_dir, "db_comparison.png"))
    plt.close()
    
    #plotting CH
    plt.figure()
    plt.plot(kmeans_k, [r["calinski_harabasz"] for r in kmeans_results], marker='o', label="KMeans")
    plt.plot(fcm_k, [r["calinski_harabasz"] for r in fcm_results], marker='s', label="FCM")
    plt.xlabel("Cluster Amount (k)")
    plt.ylabel("Calinski-Harabasz Score")
    plt.title("Calinski-Harabasz Score Comparison")
    plt.legend()
    plt.grid(True, alpha=0.3) #formatting
    plt.savefig(os.path.join(save_dir, "ch_comparison.png"))
    plt.close()
    

# ============================================================
# Main
# ============================================================

if __name__ == '__main__':

    kmeans_results = []
    fcm_results = []

    # Create plots directory so images don't clutter the project folder
    os.makedirs("plots", exist_ok=True)

    # --- Load data ---
    data, header = load_XLSX("Longotor1delta.xlsx")

    # --- Data Understanding ---
    raw_features, feature_names = data_understanding(data, header)

    # --- Preprocessing ---
    features, gene_ids, gene_names = preprocess(data, header)

    # --- Normalization (Min-Max, values between 0 and 1) ---
    features_norm, scaler = normalize(features, feature_names)

    # Convert to list-of-lists for K-Means compatibility
    features_norm_list = features_norm.tolist()

    # =========================================================
    # FUZZY C-MEANS EXPERIMENTS (Victory)
    # =========================================================

    # Experiment 1: Find optimal number of clusters via FPC
    print("=" * 60)
    print("EXPERIMENT 1: Optimal cluster count via FPC")
    print("=" * 60)
    fpcs = evaluate_fpc_vs_clusters(features_norm, m=2.0, k_range=range(2, 11))

    # Experiment 2: Run FCM with different cluster counts
    print("=" * 60)
    print("EXPERIMENT 2: FCM with varying k")
    print("=" * 60)
    for k in [2, 3, 4, 5]:
        cntr, u, labels, fpc = fuzzy_cmeans_clustering(features_norm, n_clusters=k, m=2.0)
        
        #internal metrics - Madison
        metrics = internal_fcm(features_norm, u, k)
        if metrics is not None:
            fcm_results.append({
                "method": "FCM",
                "k": k, 
                "fpc": fpc,
                "silhouette": metrics["silhouette"],
                "davies_bouldin": metrics["davies_bouldin"],
                "calinski_harabasz": metrics["calinski_harabasz"]
            })
        
        plot_fcm_clusters(features_norm, labels, cntr, feature_names, k, m=2.0)
        plot_membership_heatmap(u, k, m=2.0)
        if k == 3:
            print_cluster_members(labels, gene_ids, gene_names, k)

    # Experiment 3: Effect of fuzziness parameter m
    print("=" * 60)
    print("EXPERIMENT 3: Effect of fuzziness parameter m")
    print("=" * 60)
    m_results = evaluate_fuzziness_effect(features_norm, n_clusters=3, m_values=[1.5, 2.0, 2.5, 3.0])
    for m_val, cntr, u, labels, fpc in m_results:
        plot_fcm_clusters(features_norm, labels, cntr, feature_names, 3, m=m_val, save_prefix="fcm_m")

    # Summary table of m experiments
    print("\nFuzziness Parameter Summary (c=3):")
    print(f"  {'m':>5s}  {'FPC':>8s}  {'Cluster Sizes':>30s}")
    print("-" * 50)
    for m_val, cntr, u, labels, fpc in m_results:
        sizes = [int(np.sum(labels == i)) for i in range(3)]
        print(f"  {m_val:>5.1f}  {fpc:>8.4f}  {str(sizes):>30s}")

    # =========================================================
    # K-MEANS EXPERIMENTS (Miles / Madison)
    # =========================================================

    # Experiment 4: K-Means elbow method (SSE vs k)
    print("\n" + "=" * 60)
    print("EXPERIMENT 4: K-Means optimal cluster count (Elbow Method)")
    print("=" * 60)
    sse_l2 = evaluate_kmeans_sse(features_norm_list, k_range=range(2, 11), dist_method="L2")

    # Experiment 5: K-Means with different k values (L2 distance)
    print("=" * 60)
    print("EXPERIMENT 5: K-Means with varying k (L2)")
    print("=" * 60)
    
    for k in [2, 3, 4, 5]:
        print(f"\n--- K-Means k={k}, L2 ---")
        cp, ca, err = k_means_clustering(k, 100, 0.001, "L2", features_norm_list)
        print(f"  SSE: {err:.4f}")
        
        #internal metrics
        metrics = internal_kmeans(features_norm, ca, k)
        if metrics is not None:
            kmeans_results.append({
                "method": "KMeans",
                "k": k, 
                "sse": err,
                "silhouette": metrics["silhouette"],
                "davies_bouldin": metrics["davies_bouldin"],
                "calinski_harabasz": metrics["calinski_harabasz"]
            })
        
        cluster_counts = {}
        
        for dist_val, clust_id in ca:
            cluster_counts[clust_id] = cluster_counts.get(clust_id, 0) + 1
            
        for c in sorted(cluster_counts.keys()):
            print(f"  Cluster {c}: {cluster_counts[c]} members")
            
        plot_kmeans_clusters(features_norm, ca, cp, feature_names, k, "L2")
        
        if k == 3:
            print_kmeans_cluster_members(ca, gene_ids, gene_names, k)

    # Experiment 6: K-Means with different distance metrics
    print("\n" + "=" * 60)
    print("EXPERIMENT 6: K-Means distance metric comparison (k=3)")
    print("=" * 60)
    
    for dm in ["L1", "L2"]:
        print(f"\n--- K-Means k=3, {dm} ---")
        cp, ca, err = k_means_clustering(3, 100, 0.001, dm, features_norm_list)
        print(f"  Error: {err:.4f}")
        cluster_counts = {}
        
        for dist_val, clust_id in ca:
            cluster_counts[clust_id] = cluster_counts.get(clust_id, 0) + 1
            
        for c in sorted(cluster_counts.keys()):
            print(f"  Cluster {c}: {cluster_counts[c]} members")
            
        plot_kmeans_clusters(features_norm, ca, cp, feature_names, 3, dm)

    #comparing both models' internal metrics in list format
    print("\nComparing Internal Validation Results:")
    print (" Method     k     Silhouette     Davies-Bouldin     Calinski-Harabasz")
    print("-----------------------------------------------------------------------------------------")
            
    for r in kmeans_results:
        print(" {}  {}  {:.4f}  {:.4f} {:.4f}".format(r["method"], r["k"], r["silhouette"], r["davies_bouldin"], r["calinski_harabasz"]))
                
    for r in fcm_results:
        print(" {}  {}  {:.4f}  {:.4f} {:.4f}".format(r["method"], r["k"], r["silhouette"], r["davies_bouldin"], r["calinski_harabasz"]))

    # Plotting Internal Metrics after 
    plot_internal(kmeans_results, fcm_results)

    # end of program
    print("\n" + "=" * 60)
    print("Program Finished")
    print("=" * 60)
    print(f"\nAll plots saved to: {os.path.abspath('plots')}")